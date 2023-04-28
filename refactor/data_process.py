import openpyxl
from collections import OrderedDict
import scipy.io

import torchaudio
from torch.utils.data import Dataset
from utils import *
import re
CONSTANTS = InitializationTrain()


class CPCdata(Dataset):
    """
        Returns:
        speech [2, 96000]
        info: OrderedDict{path, score, listener, system, scene, volume, prompt}
    """

    def __init__(self, metadata, augmentation=None):
        self.metadata = metadata
        self.augmentation = augmentation
        self.resampler = torchaudio.transforms.Resample(
            orig_freq=CONSTANTS.orig_freq, new_freq=CONSTANTS.new_freq
        )

    def __len__(self):
        return len(self.metadata["path"])

    def __getitem__(self, idx):
        path = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + ".wav"
        speech, sr = torchaudio.load(path)
        # Remove the first 2 and last second (noise)
        speech = speech[:, 2 * sr: -sr]

        # Resample the audio to 16kHz sample rate
        if sr != CONSTANTS.new_freq:
            speech = self.resampler(speech)
            sr = CONSTANTS.new_freq

        # Pad or trim the audio to 6 seconds
        desired_length = sr * 6  # keep 6 seconds
        if speech.size(-1) < desired_length:
            padding = desired_length - speech.size(-1)
            speech = torch.nn.functional.pad(speech, (0, padding), "constant")
        elif speech.size(-1) > desired_length:
            speech = speech[..., :desired_length]

        if self.augmentation is not None:
            speech_augmented = self.augmentation(speech)
        else:
            speech_augmented = speech
        # speech_augmented = deepcopy(speech_augmented)

        info = OrderedDict()
        for key, value in self.metadata.items():
            info[key] = value[idx]
        # info = deepcopy(info)

        # speech size: [2, 96000]
        # path, score, listener, system, scene, volume, prompt

        return speech_augmented, info


class CPCdataMono(Dataset):
    """
        Returns:
        speech [1, 96000]
        info: OrderedDict{path, score, listener, system, scene, volume, prompt}
    """

    def __init__(self, metadata, augmentation=None):
        self.metadata = metadata
        self.augmentation = augmentation
        self.resampler = torchaudio.transforms.Resample(
            orig_freq=CONSTANTS.orig_freq, new_freq=CONSTANTS.new_freq
        )

    def __len__(self):
        return len(self.metadata["path"])

    def __getitem__(self, idx):
        path = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + "_mono.wav"
        speech, sr = torchaudio.load(path)
        speech = torch.squeeze(speech)
        # Remove the first 2 and last second (noise)
        speech = speech[2 * sr: -sr]

        # Resample the audio to 16kHz sample rate
        if sr != CONSTANTS.new_freq:
            speech = self.resampler(speech)
            sr = CONSTANTS.new_freq

        # Pad or trim the audio to 6 seconds
        desired_length = sr * 6  # keep 6 seconds
        if speech.size(-1) < desired_length:
            padding = desired_length - speech.size(-1)
            speech = torch.nn.functional.pad(speech, (0, padding), "constant")
        elif speech.size(-1) > desired_length:
            speech = speech[..., :desired_length]

        if self.augmentation is not None:
            speech_augmented = self.augmentation(speech)
        else:
            speech_augmented = speech
        # speech_augmented = deepcopy(speech_augmented)

        info = OrderedDict()
        for key, value in self.metadata.items():
            info[key] = value[idx]
        info['path'] = path
        # info = deepcopy(info)

        # speech size: [1, 96000]
        # path, score, listener, system, scene, volume, prompt

        return speech_augmented, info

class CPCdataBinaural(Dataset):
    """
        Returns:
        [speech_l speech_r] [[96000], [96000]]
        info: OrderedDict{path, score, listener, system, scene, volume, prompt}
    """

    def __init__(self, metadata, augmentation=None):
        self.metadata = metadata
        self.augmentation = augmentation
        self.resampler = torchaudio.transforms.Resample(
            orig_freq=CONSTANTS.orig_freq, new_freq=CONSTANTS.new_freq
        )

    def __len__(self):
        return len(self.metadata["path"])

    def __getitem__(self, idx):
        path_l = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + "_mono_1.wav"
        path_r = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + "_mono_2.wav"
        speech_l, sr = torchaudio.load(path_l) # [1, sz]
        speech_r, sr = torchaudio.load(path_r)
        speech_l = torch.squeeze(speech_l)
        speech_r = torch.squeeze(speech_r)
        # Remove the first 2 and last second (noise)
        speech_l = speech_l[2 * sr: -sr]
        speech_r = speech_r[2 * sr: -sr]
        # Resample the audio to 16kHz sample rate
        if sr != CONSTANTS.new_freq:
            speech_l = self.resampler(speech_l)
            speech_r = self.resampler(speech_r)
            sr = CONSTANTS.new_freq

        # Pad or trim the audio to 6 seconds
        desired_length = sr * 6  # keep 6 seconds
        if speech_l.size(-1) < desired_length:
            padding = desired_length - speech_l.size(-1)
            speech_l = torch.nn.functional.pad(speech_l, (0, padding), "constant")
            speech_r = torch.nn.functional.pad(speech_r, (0, padding), "constant")
        elif speech_l.size(-1) > desired_length:
            speech_l = speech_l[..., :desired_length]
            speech_r = speech_r[..., :desired_length]

        if self.augmentation is not None:
            speech_augmented_l = self.augmentation(speech_l)
            speech_augmented_r = self.augmentation(speech_r)
        else:
            speech_augmented_l = speech_l
            speech_augmented_r = speech_r
        # speech_augmented = deepcopy(speech_augmented)

        info = OrderedDict()
        for key, value in self.metadata.items():
            info[key] = value[idx]
        info['path'] = [path_l, path_r]
        # info = deepcopy(info)

        # speech size: [1, 96000]
        # path, score, listener, system, scene, volume, prompt

        return speech_augmented_l, speech_augmented_r, info


class ListenerInfo:
    """
        if bacthed: ListenerInfo.info is a list of dictionaries
        if signal input: ListenerInfo.info is a dictionary
    """

    def __init__(self, listener):
        # listener: ['L0231'] or ['L0231', 'L0201']
        self.workbook = openpyxl.load_workbook(CONSTANTS.LISTENER_PATH)
        self.sheet_names = self.workbook.sheetnames
        self.listener_list = self.get_listener_list()

        if type(listener) is str:
            self.info = self.get_info(listener)
        else:
            self.info = self.batch_info(listener)

    def get_listener_list(self):
        first_sheet = self.workbook[self.sheet_names[0]]
        listener_list = []
        for row in first_sheet.iter_rows(values_only=True):
            if row[0] is None:
                break
            listener_list.append(row[0][-3:])
        return listener_list  # [USER, 200, 201,...]

    def get_info(self, listener):
        """return a dictionary of listener info

        Args:
            listener (str): some thing like 'L0231'

        Returns:
            OrderedDict: A dictionary of listener info
        """
        info = OrderedDict()
        with open(CONSTANTS.AUDIOGRAM_PATH, "r") as file:
            audiogram_data = json.load(file)
        # audiogram: a 16-item list
        info["audiogram_l"] = [
            *audiogram_data[listener]["audiogram_levels_l"]
        ]
        info["audiogram_r"] = [
            *audiogram_data[listener]["audiogram_levels_r"]
        ]
        info["audiogram_cfs"] = [*audiogram_data[listener]["audiogram_cfs"]]

        return info

    def batch_info(self, listeners):
        infos = []
        for listener in listeners:
            infos.append(self.get_info(listener))
        return infos


class HurricaneData(Dataset):
    def __init__(self, state, root_dir='/home/ubuntu/elec823/hurricane', transform=None):
        self.state = state
        self.root_dir = root_dir
        self.transform = transform
        self.scores = scipy.io.loadmat(os.path.join(root_dir, 'scores.mat'))['intell']
        self.all_samples = []
        self.noise_types = {"cs":0, "ssn":1}
        self.snrs = {"snrHi":0, "snrMid":1, "snrLo":2}

        for mod_folder in os.listdir(root_dir):
            if mod_folder.startswith("."):
                continue
            mod_path = os.path.join(root_dir, mod_folder)
            if not os.path.isdir(mod_path):
                    continue
            for noise_type in os.listdir(mod_path):
                if noise_type.startswith("."):
                    continue
                noise_path = os.path.join(mod_path, noise_type)
                if not os.path.isdir(noise_path):
                    continue
                for snr in os.listdir(noise_path):
                    if snr.startswith("."):
                        continue
                    snr_path = os.path.join(noise_path, snr)
                    if not os.path.isdir(snr_path):
                            continue
                    for audio_file in os.listdir(snr_path):
                        if "mono" not in audio_file:
                            audio_path = os.path.join(snr_path, audio_file)
                            self.all_samples.append(audio_path)
        idx = 0
        val_list = []
        for i in range(0, len(self.all_samples), 180):
            val_list.extend(self.all_samples[i:i+36])
        if self.state == 'train':
            self.samples = [item for item in self.all_samples if item not in val_list]
        elif self.state == 'valid':
            self.samples = val_list

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        audio_path = self.samples[idx]
        split_str = audio_path.split('/')
        output = [split_str[-4], split_str[-3], split_str[-2], split_str[-1].split('_')[-1].split('.')[0]]
        mod = int(''.join(re.findall(r'\d+', output[0])))
        noise_type = self.noise_types[output[1]]
        snr = self.snrs[output[2]]
        utt = int(output[3])
        score = self.scores[mod-1][noise_type][snr][utt-1]
        
        waveform, sample_rate = torchaudio.load(audio_path)
        waveform = torch.sum(waveform, dim=0)
        
        # Pad or trim the audio to 6 seconds
        desired_length = sample_rate * 6  # keep 6 seconds
        if waveform.size(-1) < desired_length:
            padding = desired_length - waveform.size(-1)
            waveform = torch.nn.functional.pad(waveform, (0, padding), "constant")
        elif waveform.size(-1) > desired_length:
            waveform = waveform[..., :desired_length]
        # if self.transform:
        #     waveform = self.transform(waveform)
        mono_path = audio_path.replace('.wav', '_mono.wav')
        dict = {'path': [mono_path, mono_path]}
        return waveform, score, dict

    
class HurricaneCSData(Dataset):
    def __init__(self, state, root_dir='/home/ubuntu/elec823/hurricane', transform=None):
        self.state = state
        self.root_dir = root_dir
        self.transform = transform
        self.scores = scipy.io.loadmat(os.path.join(root_dir, 'scores.mat'))['intell']
        self.all_samples = []
        self.noise_types = {"cs":0, "ssn":1}
        self.snrs = {"snrHi":0, "snrMid":1, "snrLo":2}

        for mod_folder in os.listdir(root_dir):
            if mod_folder.startswith("."):
                continue
            ssn_path = os.path.join(root_dir, mod_folder, 'cs')
            if not os.path.isdir(ssn_path):
                continue
            for snr in os.listdir(ssn_path):
                if snr.startswith("."):
                    continue
                snr_path = os.path.join(ssn_path, snr)
                for audio_file in os.listdir(snr_path):
                    if "mono" not in audio_file:
                        audio_path = os.path.join(snr_path, audio_file)
                        self.all_samples.append(audio_path)
        idx = 0
        val_list = []
        for i in range(0, len(self.all_samples), 180):
            val_list.extend(self.all_samples[i:i+36])
        if self.state == 'train':
            self.samples = [item for item in self.all_samples if item not in val_list]
        elif self.state == 'valid':
            self.samples = val_list

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        audio_path = self.samples[idx]
        split_str = audio_path.split('/')
        output = [split_str[-4], split_str[-3], split_str[-2], split_str[-1].split('_')[-1].split('.')[0]]
        mod = int(''.join(re.findall(r'\d+', output[0])))
        noise_type = self.noise_types[output[1]]
        snr = self.snrs[output[2]]
        utt = int(output[3])
        score = self.scores[mod-1][noise_type][snr][utt-1]
        
        waveform, sample_rate = torchaudio.load(audio_path)
        waveform = torch.sum(waveform, dim=0)
        
        # Pad or trim the audio to 6 seconds
        desired_length = sample_rate * 6  # keep 6 seconds
        if waveform.size(-1) < desired_length:
            padding = desired_length - waveform.size(-1)
            waveform = torch.nn.functional.pad(waveform, (0, padding), "constant")
        elif waveform.size(-1) > desired_length:
            waveform = waveform[..., :desired_length]
        # if self.transform:
        #     waveform = self.transform(waveform)
        mono_path = audio_path.replace('.wav', '_mono.wav')
        dict = {'path': [mono_path, mono_path]}
        return waveform, score, dict
    

class HurricaneSSNData(Dataset):
    def __init__(self, state, root_dir='/home/ubuntu/elec823/hurricane', transform=None):
        self.state = state
        self.root_dir = root_dir
        self.transform = transform
        self.scores = scipy.io.loadmat(os.path.join(root_dir, 'scores.mat'))['intell']
        self.all_samples = []
        self.noise_types = {"cs":0, "ssn":1}
        self.snrs = {"snrHi":0, "snrMid":1, "snrLo":2}

        for mod_folder in os.listdir(root_dir):
            if mod_folder.startswith("."):
                continue
            ssn_path = os.path.join(root_dir, mod_folder, 'ssn')
            if not os.path.isdir(ssn_path):
                continue
            for snr in os.listdir(ssn_path):
                if snr.startswith("."):
                    continue
                snr_path = os.path.join(ssn_path, snr)
                for audio_file in os.listdir(snr_path):
                    if "mono" not in audio_file:
                        audio_path = os.path.join(snr_path, audio_file)
                        self.all_samples.append(audio_path)
        idx = 0
        val_list = []
        for i in range(0, len(self.all_samples), 180):
            val_list.extend(self.all_samples[i:i+36])
        if self.state == 'train':
            self.samples = [item for item in self.all_samples if item not in val_list]
        elif self.state == 'valid':
            self.samples = val_list

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        audio_path = self.samples[idx]
        split_str = audio_path.split('/')
        output = [split_str[-4], split_str[-3], split_str[-2], split_str[-1].split('_')[-1].split('.')[0]]
        mod = int(''.join(re.findall(r'\d+', output[0])))
        noise_type = self.noise_types[output[1]]
        snr = self.snrs[output[2]]
        utt = int(output[3])
        score = self.scores[mod-1][noise_type][snr][utt-1]
        
        waveform, sample_rate = torchaudio.load(audio_path)
        waveform = torch.sum(waveform, dim=0)
        
        # Pad or trim the audio to 6 seconds
        desired_length = sample_rate * 6  # keep 6 seconds
        if waveform.size(-1) < desired_length:
            padding = desired_length - waveform.size(-1)
            waveform = torch.nn.functional.pad(waveform, (0, padding), "constant")
        elif waveform.size(-1) > desired_length:
            waveform = waveform[..., :desired_length]
        # if self.transform:
        #     waveform = self.transform(waveform)
        mono_path = audio_path.replace('.wav', '_mono.wav')
        dict = {'path': [mono_path, mono_path]}
        return waveform, score, dict