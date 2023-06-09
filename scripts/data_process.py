import openpyxl
import json
from collections import OrderedDict

import torchaudio
from torch.utils.data import Dataset
from utils import *

CONSTANTS = InitializationTrain(verbose=True)

class CPCdata(Dataset):
    """
        Returns:
        speech [2, 96000]
        info: OrderedDict{path, score, listener, system, scene, volume, prompt}
    """

    def __init__(self, metadata, augmentation=None):
        self.metadata = metadata
        self.augmentation = augmentation
        self.resampler = torchaudio.transforms.Resample(
            orig_freq=CONSTANTS.orig_freq, new_freq=CONSTANTS.new_freq
        )

    def __len__(self):
        return len(self.metadata["path"])

    def __getitem__(self, idx):
        path = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + ".wav"
        speech, sr = torchaudio.load(path)
        # Remove the first 2 and last second (noise)
        speech = speech[:, 2 * sr : -sr]

        # Resample the audio to 16kHz sample rate
        if sr != CONSTANTS.new_freq:
            speech = self.resampler(speech)
            sr = CONSTANTS.new_freq

        # Pad or trim the audio to 6 seconds
        desired_length = sr * 6  # keep 6 seconds
        if speech.size(-1) < desired_length:
            padding = desired_length - speech.size(-1)
            speech = torch.nn.functional.pad(speech, (0, padding), "constant")
        elif speech.size(-1) > desired_length:
            speech = speech[..., :desired_length]

        if self.augmentation is not None:
            speech_augmented = self.augmentation(speech)
        else:
            speech_augmented = speech
        # speech_augmented = deepcopy(speech_augmented)

        info = OrderedDict()
        for key, value in self.metadata.items():
            info[key] = value[idx]
        # info = deepcopy(info)

        # speech size: [2, 96000]
        # path, score, listener, system, scene, volume, prompt

        return speech_augmented, info

class CPCdataMono(Dataset):
    """
        Returns:
        speech [2, 96000]
        info: OrderedDict{path, score, listener, system, scene, volume, prompt}
    """

    def __init__(self, metadata, augmentation=None):
        self.metadata = metadata
        self.augmentation = augmentation
        self.resampler = torchaudio.transforms.Resample(
            orig_freq=CONSTANTS.orig_freq, new_freq=CONSTANTS.new_freq
        )

    def __len__(self):
        return len(self.metadata["path"])

    def __getitem__(self, idx):
        path = CONSTANTS.DATA_PATH + self.metadata["path"][idx] + "_mono.wav"
        speech, sr = torchaudio.load(path)
        speech = torch.squeeze(speech)
        # Remove the first 2 and last second (noise)
        speech = speech[2 * sr : -sr]

        # Resample the audio to 16kHz sample rate
        if sr != CONSTANTS.new_freq:
            speech = self.resampler(speech)
            sr = CONSTANTS.new_freq

        # Pad or trim the audio to 6 seconds
        desired_length = sr * 6  # keep 6 seconds
        if speech.size(-1) < desired_length:
            padding = desired_length - speech.size(-1)
            speech = torch.nn.functional.pad(speech, (0, padding), "constant")
        elif speech.size(-1) > desired_length:
            speech = speech[..., :desired_length]

        if self.augmentation is not None:
            speech_augmented = self.augmentation(speech)
        else:
            speech_augmented = speech
        # speech_augmented = deepcopy(speech_augmented)

        info = OrderedDict()
        for key, value in self.metadata.items():
            info[key] = value[idx]
        # info = deepcopy(info)

        # speech size: [2, 96000]
        # path, score, listener, system, scene, volume, prompt

        return speech_augmented, info


class ListenerInfo:
    """
        if bacthed: ListenerInfo.info is a list of dictionaries
        if signal input: ListenerInfo.info is a dictionary
    """

    def __init__(self, listener):
        # listener: ['L0231'] or ['L0231', 'L0201']
        self.workbook = openpyxl.load_workbook(CONSTANTS.LISTENER_PATH)
        self.sheet_names = self.workbook.sheetnames
        self.listener_list = self.get_listener_list()

        if type(listener) is str:
            self.info = self.get_info(listener)
        else:
            self.info = self.batch_info(listener)

    def get_listener_list(self):
        first_sheet = self.workbook[self.sheet_names[0]]
        listener_list = []
        for row in first_sheet.iter_rows(values_only=True):
            if row[0] is None:
                break
            listener_list.append(row[0][-3:])
        return listener_list  # [USER, 200, 201,...]

    def get_info(self, listener):
        """return a dictionary of listener info

        Args:
            listener (str): some thing like 'L0231'

        Returns:
            OrderedDict: A dictionary of listener info
        """
        info = OrderedDict()
        with open(CONSTANTS.AUDIOGRAM_PATH, "r") as file:
            audiogram_data = json.load(file)
        # audiogram: a 16-item list
        info["audiogram"] = [
            *audiogram_data[listener]["audiogram_levels_l"],
            *audiogram_data[listener]["audiogram_levels_r"],
        ]

        return info

    def batch_info(self, listeners):
        infos = []
        for listener in listeners:
            infos.append(self.get_info(listener))
        return infos